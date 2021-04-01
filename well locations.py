# -*- coding: utf-8 -*-
"""
Created on Wed Jan 13 12:35:01 2021

@author: Alexa
"""

import sys
import os
import openpyxl as xl
import numpy as np
import matplotlib.pyplot as plt
import math

DATADIR = "D:/Induced Earthquakes/dataInput"
WKDIR = "D:/Induced Earthquakes"
TESTFILE = ["Well_Production_API_03701984.xlsx"]
TESTFILE2 = ["Well_Injection_API_23701291.xlsx"]
STARTYEAR = 1980
ENDYEAR = 2017
    
files = os.listdir(DATADIR)
os.chdir(WKDIR)

production = [x for x in files if "Production" in x]
injection = [x for x in files if "Injection" in x]

print(len(production))
print(len(injection))
print(TESTFILE[0][20:28])
print(TESTFILE2[0][19:27])
#if 'WellsAPInumber.dat' in files: files.remove('WellsAPInumber.dat')
#print(files)

all_production = {}
all_locations = {}

for well in production:
    workbook = xl.load_workbook(filename=f"{DATADIR}/{well}", read_only=True, data_only=True)
    sheet = workbook.active
    well_dates = {}
    for yields in sheet.iter_rows(min_row = 5, min_col = 2, max_col = 5, values_only=True):
        yields = [0 if x is None else x for x in yields]
        well_dates[yields[0]] = [yields[1], yields[2], yields[3]]
    for location in sheet.iter_rows(min_row = 2, min_col = 14, max_row = 2, max_col = 15, values_only=True):
        all_locations[well] = [location[0], location[1]]
        
    all_production[well] = well_dates
    
for well in injection:
    workbook = xl.load_workbook(filename=f"{DATADIR}/{well}", read_only=True, data_only=True)
    sheet = workbook.active
    well_dates = {}
    for yields in sheet.iter_rows(min_row = 5, min_col = 2, max_col = 4, values_only=True):
        yields = [0 if x is None else x for x in yields]
        well_dates[yields[0]] = [yields[1], yields[2]]
    for location in sheet.iter_rows(min_row = 2, min_col = 14, max_row = 2, max_col = 15, values_only=True):
        all_locations[well] = [location[0], location[1]]
    all_production[well] = well_dates

#print(all_production)


years = [*range(STARTYEAR, ENDYEAR + 1)]
year_values = [f"{year} Total" for year in years]

all_years = {}
culmative_yearly_production = []
for year in year_values: culmative_yearly_production.append(0)
culmative_yearly_production = np.asarray(culmative_yearly_production)
type_yearly_production = np.asarray([[0,0,0]])
for year in year_values: type_yearly_production = np.vstack([type_yearly_production, [[0,0,0]]])
type_yearly_production = np.delete(type_yearly_production, (0), axis=0)
#print(type_yearly_production)
culmative_yearly_injection = []
for year in year_values: culmative_yearly_injection.append(0)


#print(year_values)
 
for well in production:
    yearly_total_pumped = []
    yearly_typed_pumped = np.asarray([0,0,0])
    for year in year_values:
        try:
            yearly_total_pumped.append(sum(all_production[well][year][0:1]))
            yearly_typed_pumped = np.vstack([yearly_typed_pumped, [all_production[well][year]]])
        except:
            yearly_total_pumped.append(0)
            yearly_typed_pumped = np.vstack([yearly_typed_pumped, [[0,0,0]]])
    yearly_typed_pumped = np.delete(yearly_typed_pumped, (0), axis=0)
    #print(yearly_total_pumped)
    culmative_yearly_production += np.asarray(yearly_total_pumped)
    #print(type_yearly_production)
    #print(yearly_typed_pumped)
    type_yearly_production = type_yearly_production + yearly_typed_pumped
    all_years[well] = yearly_total_pumped

for well in injection:
    yearly_total_injected = []
    for year in year_values:
        try:
            yearly_total_injected.append(sum(all_production[well][year][0]))
        except:
            yearly_total_injected.append(0)
    culmative_yearly_injection += np.asarray(yearly_total_injected)
    all_years[well] = yearly_total_injected

well_injection = {}
for year in year_values:
    yearly_values = []
    for well in injection:
        if float(all_locations[well][0]) != 0 and float(all_locations[well][1]) != 0:
            try:
                    yearly_values.append(math.log(float(all_production[well][year][0]), 10))
            except: 
                yearly_values.append(0)
    well_injection[year] = yearly_values
    
well_net = {}
for year in year_values:
    yearly_values = []
    for well in injection:
         if float(all_locations[well][0]) != 0 and float(all_locations[well][1]) != 0:
             for p_well in production:
                 if well[19:27] == p_well[20:28]:
                     try:
                         plus = float(all_production[well][year][0])
                     except: 
                         plus = 0
                     try:
                         minus = sum(all_production[p_well][year][0:1])
                     except:
                         minus = 0
                     net = plus - minus
                     if net > 0:
                         yearly_values.append(math.log(net))
                     if net == 0:
                         yearly_values.append(0)
                     else:
                         yearly_values.append(-1 * math.log(-1 * net))
    well_net[year] = yearly_values

well_production = {}
oil_production = {}
for year in year_values:
    yearly_values = []
    yearly_oil = []
    for well in production:
        if float(all_locations[well][0]) != 0 and float(all_locations[well][1]) != 0:
            try:
                    yearly_values.append(math.log(sum(all_production[well][year][0:1]),10))
                    yearly_oil.append(all_production[well][year][0])
            except: 
                yearly_values.append(0)
    well_production[year] = yearly_values

#print(culmative_yearly_production)
print(type_yearly_production)

culmative_yearly_production = culmative_yearly_production * (10 ** -9)
type_yearly_production = type_yearly_production * (10 ** -9)

i_latitude = []
i_longitude = []
for well in injection:
    if float(all_locations[well][0]) != 0 and float(all_locations[well][1]) != 0:
        i_latitude.append(float(all_locations[well][0]))
        i_longitude.append(float(all_locations[well][1]))
        
        

p_latitude = []
p_longitude = []
for well in production:
    if float(all_locations[well][0]) != 0 and float(all_locations[well][1]) != 0:
        p_latitude.append(float(all_locations[well][0]))
        p_longitude.append(float(all_locations[well][1]))
        
DATADIR = "D:/Induced Earthquakes/Wilmington Oil Earthquakes"
WKDIR = "D:/Induced Earthquakes"
STARTYEAR = 1980
ENDYEAR = 2017
earthquakes = open(f"{DATADIR}/SearchResults.txt", "r")
lines = earthquakes.readlines()


earthquakes = []

cats = lines[2].split()

print(cats)

for line in lines:
    row = line.split()
    data = {}
    if len(row) == 13 and row[2] != 'ET':
        for i in range(len(cats)):
            try:
                value = float(row[i])
                data[cats[i]] = value
            except:
                data[cats[i]] = row[i]
        earthquakes.append(data)
    
print(earthquakes)

latitude = []
longitude = []
magnitude = []
e_years = []

for earthquake in earthquakes:
    latitude.append(earthquake['LAT'])
    longitude.append(earthquake['LON'])
    magnitude.append(earthquake['MAG'])
    year = earthquake['#YYY/MM/DD'].split("/")[0]
    e_years.append(int(year))
    
earthquakes_by_year = {}

for i in range(len(years)):
    year = years[i]
    vyear = year_values[i]
    latitudes = []
    longitudes = []
    magnitudes = []
    for i in range(len(e_years)):
        if int(year) == int(e_years[i]):
            latitudes.append(latitude[i])
            longitudes.append(longitude[i])
            magnitudes.append(magnitude[i])
    year_of_earthquakes = {}
    year_of_earthquakes["latitude"] = latitudes
    year_of_earthquakes["longitude"] = longitudes
    year_of_earthquakes["magnitude"] = magnitudes
    
    earthquakes_by_year[vyear] = year_of_earthquakes
    
for year in year_values:
    plt.scatter(i_longitude, i_latitude, c=well_injection[year], cmap='Blues', alpha = .3)
    #vmin=0, vmax=5*10**6
    plt.title(f"Location of Oil Wells In Wilmington Oil Field in {year} with Injected Water")
    plt.xlabel("Latitude")
    plt.ylabel("Longitude")
    plt.colorbar()
    plt.clim(5, 8)
    plt.scatter(earthquakes_by_year[year]["longitude"], earthquakes_by_year[year]["latitude"], c="Black")
    plt.ylim(33.71,33.82)
    plt.xlim(-118.33,-118.13)
    plt.show()  
    
for year in year_values:
    plt.scatter(p_longitude, p_latitude, c=well_production[year], cmap='Reds', alpha = .3)
    #vmin=0, vmax=5*10**6
    plt.title(f"Location of Oil Wells In Wilmington Oil Field in {year} with Removed Water")
    plt.xlabel("Latitude")
    plt.ylabel("Longitude")
    plt.colorbar()
    plt.clim(3,7)
    plt.ylim(33.71,33.82)
    plt.xlim(-118.33,-118.13)
    plt.scatter(earthquakes_by_year[year]["longitude"], earthquakes_by_year[year]["latitude"], c="Black")
    plt.show()
    
for year in year_values:
    plt.scatter(i_longitude, i_latitude, c=well_net[year], cmap='RdBu', alpha = .3)
    #vmin=0, vmax=5*10**6
    plt.title(f"Location of Oil Wells In Wilmington Oil Field in {year} with Injected Water")
    plt.xlabel("Latitude")
    plt.ylabel("Longitude")
    plt.colorbar()
    plt.clim(-7, 7)
    plt.scatter(earthquakes_by_year[year]["longitude"], earthquakes_by_year[year]["latitude"], c="Black")
    plt.ylim(33.71,33.82)
    plt.xlim(-118.33,-118.13)
    plt.show()  
    
for year in year_values:
    plt.scatter(earthquakes_by_year[year]["longitude"], earthquakes_by_year[year]["latitude"], c="Black")
    plt.title(f"Location of Earthquakes in Wilmington Oil Field in {year} with Injected and Removed Water")
    plt.xlabel("Latitude")
    plt.ylabel("Longitude")
    plt.ylim(33.71,33.82)
    plt.xlim(-118.33,-118.13)
    plt.show() 
    
print(len(i_longitude))
print(len(p_longitude))

for year in year_values:
    largest_injection = 0
    for value in well_injection[year]:
        if value > largest_injection: largest_injection = value


        
      
    
    

    