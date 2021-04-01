# -*- coding: utf-8 -*-
"""
Created on Tue Jan 12 11:19:22 2021

@author: Alexa
"""

import sys
import os
import openpyxl as xl
import numpy as np
import matplotlib.pyplot as plt

DATADIR = "D:/Induced Earthquakes/dataInput"
WKDIR = "D:/Induced Earthquakes"
TESTFILE = ["Well_Production_API_03701984.xlsx"]
TESTFILE2 = ["Well_Injection_API_23701291.xlsx"]
STARTYEAR = 1980
ENDYEAR = 2017

files = os.listdir(DATADIR)
os.chdir(WKDIR)

production = [x for x in files if "Production" in x]
injection = [x for x in files if "Injection" in x]

#if 'WellsAPInumber.dat' in files: files.remove('WellsAPInumber.dat')
#print(files)

all_production = {}

for well in production:
    workbook = xl.load_workbook(filename=f"{DATADIR}/{well}", read_only=True, data_only=True)
    sheet = workbook.active
    well_dates = {}
    for yields in sheet.iter_rows(min_row = 5, min_col = 2, max_col = 5, values_only=True):
        yields = [0 if x is None else x for x in yields]
        well_dates[yields[0]] = [yields[1], yields[2], yields[3]]
    all_production[well] = well_dates
    
for well in injection:
    workbook = xl.load_workbook(filename=f"{DATADIR}/{well}", read_only=True, data_only=True)
    sheet = workbook.active
    well_dates = {}
    for yields in sheet.iter_rows(min_row = 5, min_col = 2, max_col = 4, values_only=True):
        yields = [0 if x is None else x for x in yields]
        well_dates[yields[0]] = [yields[1], yields[2]]
    all_production[well] = well_dates

#print(all_production)

years = range(STARTYEAR, ENDYEAR + 1)
year_values = [f"{year} Total" for year in years]

all_years = {}
culmative_yearly_production = []
for year in year_values: culmative_yearly_production.append(0)
culmative_yearly_production = np.asarray(culmative_yearly_production)
type_yearly_production = np.asarray([[0,0,0]])
for year in year_values: type_yearly_production = np.vstack([type_yearly_production, [[0,0,0]]])
type_yearly_production = np.delete(type_yearly_production, (0), axis=0)
#print(type_yearly_production)
culmative_yearly_injection = []
for year in year_values: culmative_yearly_injection.append(0)


#print(year_values)
 
for well in production:
    yearly_total_pumped = []
    yearly_typed_pumped = np.asarray([0,0,0])
    for year in year_values:
        try:
            yearly_total_pumped.append(sum(all_production[well][year][0:1]))
            yearly_typed_pumped = np.vstack([yearly_typed_pumped, [all_production[well][year]]])
        except:
            yearly_total_pumped.append(0)
            yearly_typed_pumped = np.vstack([yearly_typed_pumped, [[0,0,0]]])
    yearly_typed_pumped = np.delete(yearly_typed_pumped, (0), axis=0)
    #print(yearly_total_pumped)
    culmative_yearly_production += np.asarray(yearly_total_pumped)
    #print(type_yearly_production)
    #print(yearly_typed_pumped)
    type_yearly_production = type_yearly_production + yearly_typed_pumped
    all_years[well] = yearly_total_pumped
    
for well in injection:
    yearly_total_injected = []
    for year in year_values:
        try:
            yearly_total_injected.append(sum(all_production[well][year][0]))
        except:
            yearly_total_injected.append(0)
    culmative_yearly_injection += np.asarray(yearly_total_injected)
    all_years[well] = yearly_total_injected

#print(culmative_yearly_production)
print(type_yearly_production)

culmative_yearly_production = culmative_yearly_production * (10 ** -9)
type_yearly_production = type_yearly_production * (10 ** -9)

culmative_yearly_injection = culmative_yearly_injection * (10**-9)

net_yearly_removed = culmative_yearly_production - culmative_yearly_injection


DATADIR = "D:/Induced Earthquakes/Wilmington Oil Earthquakes"
WKDIR = "D:/Induced Earthquakes"
STARTYEAR = 1980
ENDYEAR = 2018

earthquakes = open(f"{DATADIR}/SearchResults.txt", "r")
lines = earthquakes.readlines()

earthquakes = []

cats = lines[2].split()

print(cats)

for line in lines:
    row = line.split()
    data = {}
    if len(row) == 13 and row[2] != 'ET':
        for i in range(len(cats)):
            try:
                value = float(row[i])
                data[cats[i]] = value
            except:
                data[cats[i]] = row[i]
        earthquakes.append(data)
    
print(earthquakes)

latitude = []
longitude = []
magnitude = []
e_years = []

for earthquake in earthquakes:
    latitude.append(earthquake['LAT'])
    longitude.append(earthquake['LON'])
    magnitude.append(earthquake['MAG'])
    year = earthquake['#YYY/MM/DD'].split("/")[0]
    e_years.append(int(year))
    
occurances = counter(year)
 
print(len(magnitude))
print(len(e_years))
plt.scatter(e_years, magnitude)
plt.plot(years, net_yearly_removed)
plt.show()

fix, ax = plt.subplots()
ax.plot(years, net_yearly_removed)
ax.set_xlabel("years")
ax.set_ylabel("Net Liquid Pumped By Year")

ax2 = ax.twinx()
ax2.scatter(e_years, magnitude,color="red",marker="o")
ax2.set_ylabel("magnitude",color="blue",fontsize=14)
plt.show()









